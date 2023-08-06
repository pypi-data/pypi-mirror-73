import json
from itertools import islice
from pprint import pprint

import openpyxl


class ExcelConverter(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.excel_json = {
            "projectproperties": {
                "name": "Touchless Project",
                "version": "1.0.0"
            },
            "pages": []
        }
        self.workbook = openpyxl.load_workbook(filename=self.file_name)

    def meta_row_obj(self, repeats_info):
        meta_info = {
            "title": next(filter(lambda x: x['header'] == 'Label', repeats_info)).get('val'),
            "description": "",
            "properties": {
                "join": str(next(filter(lambda x: x['header'] == 'Join', repeats_info)).get('val'))
            }
        }
        return meta_info

    def row_to_obj(self, row_data):
        col_obj  = {
            "cols": []
        }
        for col_groups in row_data:
            col_json = {
                "type": None,
                "properties": {
                }
            }
            for col in col_groups:
                header_val = col['header']
                col_val = col['val']

                if header_val == "Control Type":
                    col_json['type'] = col_val.lower()
                elif header_val == 'Join':
                    col_json['properties']['join'] = str(col_val)
                elif header_val == 'Label':
                    col_json['properties']['label'] = col_val
                else:
                    y_n_translation = {"Y": "yes", "N": "no"}.get(col_val.upper(), None)
                    col_val = y_n_translation if y_n_translation is not None else col_val
                    col_json['properties'][header_val.lower().replace(' ', '-')] = col_val

            col_obj["cols"].append(col_json)
        
        return col_obj

    def convert(self):
        sheets = self.workbook.sheetnames
        for sheet_name in sheets:
            excel_sheet = self.workbook[sheet_name]
            sheet_json = {
                "rows": []
            }
            row_json = {
            }
            for row in islice(excel_sheet.iter_rows(), 1, None):
                temp_row = [{'cell': c, 'header': excel_sheet.cell(1, c.column).value, 'val':c.value} for c in row]
                temp_row = list(filter(lambda x: x['val'] is not None, temp_row))
                slices = [idx for idx, vals in enumerate(temp_row) if vals['header'] == temp_row[0]['header']] + [len(temp_row)]
                repeats = [temp_row[slices[i]:slices[(i + 1) % len(slices)]] for i in range(len(slices) - 1)]

                # pprint(repeats)
                if len(repeats) == 1 and next(map(lambda x: x['val'] == "Header", repeats[0])):
                    update_val = self.meta_row_obj(repeats[0])
                    sheet_json.update(update_val)
                else:
                    if repeats:
                        gen_controls = self.row_to_obj(repeats)
                        sheet_json["rows"].append(gen_controls)
                # for row_repeats in repeats:
            self.excel_json["pages"].append(sheet_json)
