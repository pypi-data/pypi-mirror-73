"""
Utils for working with openpyxl.Workbook
"""

from io import BytesIO
from typing import Sequence, Dict

from openpyxl import Workbook, load_workbook

from docci.file import FileAttachment


def xlsx_to_file(xlsx: Workbook, name: str) -> FileAttachment:
    """
    Convert openpyxl.Workbook to FileAttachment
    """
    return FileAttachment(name, xlsx_to_bytes(xlsx))


def xlsx_to_bytes(xlsx: Workbook) -> bytes:
    """
    Convert openpyxl.Workbook to bytes
    """
    excel_stream = BytesIO()
    xlsx.save(excel_stream)
    return excel_stream.getvalue()


def dicts_to_xlsx(dicts: Sequence[Dict], headers: Sequence[str] = None) -> Workbook:
    """
    Create openpyxl.Workbook with rows of {dicts} values.

    :param dicts: List of dicts to insert
    :param headers: List of headers if None dict keys would be used.
    :return: openpyxl.Workbook
    """
    headers = headers or tuple(dicts[0].keys())

    xlsx = Workbook()
    sheet = xlsx.active

    sheet.append(headers)
    for dict_ in dicts:
        sheet.append(tuple(dict_.values()))

    return xlsx


def xlsx_from_bytes(bytes_: bytes) -> Workbook:
    """Create xlsx from bytes."""
    return load_workbook(BytesIO(bytes_))


def xlsx_from_file(file: FileAttachment) -> Workbook:
    """Create xlsx from FileAttachment"""
    return xlsx_from_bytes(file.content)
